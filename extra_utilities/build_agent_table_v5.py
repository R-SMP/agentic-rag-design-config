"""Build the v5 version of the agent-description table.

Updates the original Agent-Database_consolidated.xlsx layout to reflect
every behavioural change shipped through this conversation, and adds a
new column "Additional RAG database access (suggested) + reason" for
each agent.

Output: C:/Users/vince/OneDrive/Desktop/MT/Meetings/04.21/
        Agents_list_wdatabase_connections/Agent-Database_v5.xlsx
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUT_PATH = Path(
    r"C:\Users\vince\OneDrive\Desktop\MT\Meetings\04.21"
    r"\Agents_list_wdatabase_connections\Agent-Database_v5.xlsx"
)

HEADERS = [
    "Goal of task/step",
    "Agent / Sub-Agent",
    "Has Message history in its context?",
    "Inputs",
    "Outputs",
    "Database elements retrievable that can be added to context "
    "before calling the agent",
    "Additional RAG database access (suggested) + short reason",
]


def b(*lines: str) -> str:
    """Join lines into a single cell value with newlines."""
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Row data (one tuple per agent)
# Order matches the original spreadsheet's row order.
# ---------------------------------------------------------------------------

ROWS = [
    # ============== 1. Planner ==============
    (
        b(
            "Decides what to do at every step of the workflow:",
            "  - shapes the plan (Problem / Solution / Sequence)",
            "  - emits an attempt-budget audit line (HARD RULE 12)",
            "  - is called on every escalation; no fixed cadence",
            "  - knows every agent + every tool generically (no per-agent details)",
            "  - produces a plan with three escalation framings: locked-value collision / "
            "out-of-qualitative-levers / both",
            "Acts as the actual 'brain' of the workflow — Orchestrator originates nothing,",
            "Planner originates qualitative directives.",
            "(RAG NOT YET WIRED) Per step, decides whether/what to retrieve from DB.",
            "Decisions must be engineering-grounded and respect the context window.",
        ),
        "Planner",
        b(
            "Y",
            "Own self.messages persists across hand-offs.",
            "Includes own system prompt as the first message.",
        ),
        b(
            "User inputs file location",
            "If present:",
            "  - Current system conditions",
            "  - Error messages",
            "  - Notes from other agents",
            "  - Human feedback",
            "Optional read tools: read_user_queries, read_image_notes, "
            "read_agent_history, list/read attempts.",
        ),
        b(
            "Plan (Problem / Solution / Sequence) — chosen agent to call next "
            "and what to check (target may be Planner itself) (Semantic)",
            "Specific notes for named downstream agents",
            "(Future) Per-step RAG retrieval directives",
        ),
        b(
            "Problem formulation (Semantic)",
            "Solution to the problem (Semantic)",
            "Failure Log (Semantic)",
            "Design Intent and Functional Requirements (Semantic)",
            "Plan in Steps (also showing which tools to call and what to check) (Semantic)",
        ),
        b(
            "• Plan-to-outcome history (Semantic): correlate prior plan shapes with "
            "approval / escalation outcomes; avoids replaying plans that historically "
            "failed.",
            "",
            "• Recovery-pattern library (Semantic): catalogued routes for known "
            "agent-side bugs (DCIC argument-shape, UII parenthetical-noise) so the "
            "Planner doesn't have to re-discover them from first principles.",
            "",
            "• Three-framing examples library (Semantic): worked examples of HARD "
            "RULE 12's locked-value-collision / out-of-levers / both framings beyond "
            "the in-prompt examples — improves escalation phrasing.",
        ),
    ),

    # ============== 2. User Input Inspector ==============
    (
        b(
            "Reads user inputs (text + images + image notes) and produces a structured "
            "extraction.",
            "Three sections in extracted_inputs.txt:",
            "  1. QUANTITATIVE INPUTS — canonical configurator parameter names "
            "(camelCase: bladeCount, impellerRadius, …); real-world entries flagged.",
            "  2. QUALITATIVE DESCRIPTIONS — image character + sketch judgement.",
            "  3. DESIGN INTENT AND FUNCTIONAL REQUIREMENTS.",
            "Adds '(unlocked by user)' annotation when the user explicitly authorises a "
            "value to vary.",
            "Determines if the reference image qualifies as a sketch (counts precise; "
            "wobble = drawing artifact).",
        ),
        "User Input Inspector (UII)",
        b(
            "Y",
            "Own self.messages persists across hand-offs.",
            "Image bytes stripped at on_operation_end when "
            "KEEP_IMAGES_IN_CONTEXT=False; paired path-text labels retained.",
        ),
        b(
            "User Inputs (inputs/ + inputs/input_images/<name>.png|.jpg|.jpeg + "
            "<name>_note.txt pairs)",
            "Hand-off message from Planner",
            "If present: Planner notes for input analysis",
        ),
        b(
            "1. DC inputs quantitatively extrapolated from user inputs",
            "   (under canonical camelCase parameter names)",
            "2. Qualitative descriptions of design inputs extracted from user inputs",
            "3. Design Intent and Functional Requirements",
            "All written into extracted_inputs.txt",
        ),
        b(
            "Design Intent and Functional Requirements (Semantic)",
            "User-defined inputs (Floats and Ints) (Quantitative)",
            "User-defined inputs (2D files) (Quantitative) (Currently not implemented)",
            "User-defined inputs (3D files) (Quantitative) (Currently not implemented)",
            "User-defined inputs (image) (Semantic) (Currently not implemented)",
            "Input Parameters for Design Configurator (Floats and Ints) (Quantitative)",
            "Input Parameters for Design Configurator (2D files) (Quantitative) "
            "(Currently not implemented)",
            "Input Parameters for Design Configurator (3D files) (Quantitative) "
            "(Currently not implemented)",
            "Input Parameters for Design Configurator (images) (Semantic) "
            "(Currently not implemented)",
        ),
        b(
            "• Sketch-vs-render judgement examples (Semantic + image): past extractions "
            "labelled sketch / render / hybrid with ground truth; sharpens the "
            "sketch-handling decision.",
            "",
            "• Unit-conversion pattern library (Semantic): worked examples of "
            "real-world callouts converted into configurator units (mm → % of chord, "
            "fraction of radius, etc.) — prevents conversion mistakes when downstream "
            "chord values change.",
            "",
            "• Annotation-callout corpus (Semantic + image): how arrow-and-label "
            "sketch annotations on rendered images have been correctly read in past "
            "sessions; covers the hybrid case the v5 paired-note workflow targets.",
        ),
    ),

    # ============== 3. DC Input Creator (DCIC) ==============
    (
        b(
            "Reads extracted_inputs.txt and authors the complete 17-key parameters.json "
            "for the current attempt folder.",
            "  - User-locked values: copy verbatim.",
            "  - Unlocked values: choose within range, with rationale tied to the "
            "qualitative cues.",
            "  - Calls write_parameters({attempt_dir, parameters: {…}}) — both "
            "arguments required.",
            "Recovery cycles: when DCOI returns REVISE, DCIC opens a NEW attempt and "
            "writes the next iteration. Must preserve user-locked values and any "
            "real-world callout conversions across iterations.",
            "Known issue (v5 mini-class models): may drop the 'parameters' arg on "
            "first call — full-class models self-correct on the next call.",
        ),
        "DC Input Creator (DCIC)",
        b(
            "Y",
            "Own self.messages persists across hand-offs.",
            "Image bytes stripped at on_operation_end (image-loading is available "
            "but rarely needed since the extraction already captures intent).",
        ),
        b(
            "Hand-off from User Input Inspector",
            "extracted_inputs.txt",
            "If present: Planner notes for DCIC (e.g. recovery directives like "
            "'reduce hub prominence', 'broaden mid-station chord')",
            "Optional: prior parameters.json via read_attempt(n, 'parameters.json')",
        ),
        b(
            "Complete 17-key parameters.json written into the current attempt folder, "
            "in canonical camelCase order.",
            "Hand-off to DC Input Inspector documenting:",
            "  - which values are user-locked (copied verbatim from extraction)",
            "  - which values are DCIC-chosen + rationale tied to qualitative cues",
            "  - any conversion identities applied (e.g. impellerRadius = diameter/2)",
        ),
        b(
            "Design Intent and Functional Requirements (Semantic)",
            "Input Parameters for Design Configurator (Floats and Ints) (Quantitative)",
            "Input Parameters for Design Configurator (2D files) (Quantitative) "
            "(Currently not implemented)",
            "Input Parameters for Design Configurator (3D files) (Quantitative) "
            "(Currently not implemented)",
            "Input Parameters for Design Configurator (images) (Semantic) "
            "(Currently not implemented)",
        ),
        b(
            "• Approved-parameter-sets library (Quantitative + Semantic): all "
            "parameters.json from approved attempts indexed by qualitative description "
            "of the request, so similar requests can seed from a vetted set instead of "
            "starting from pure assumption.",
            "",
            "• Parameter-correlation priors (Quantitative): pairwise statistics "
            "(e.g. 'when middlePos is 0.55, middleChord is typically 22-26 mm') so "
            "unlocked-value choices are statistically grounded, not arbitrary.",
            "",
            "• Conversion-anchor library (Semantic + Quantitative): past iterations "
            "where a real-world callout was preserved across chord changes — directly "
            "addresses the 2.28 mm callout drop seen in the recent recovery loop.",
        ),
    ),

    # ============== 4. DC Input Inspector (DCII) ==============
    (
        b(
            "Validates the newly-written parameters.json against four checks:",
            "  1. Per-parameter range bounds (every parameter individually).",
            "  2. Consistency with extracted_inputs.txt — QUANTITATIVE INPUTS must "
            "match verbatim; '(unlocked by user)' is the ONLY signal allowing override; "
            "DESIGN INTENT is informational.",
            "  3. Engineering hard blockers (innerThickness > 0, outerThickness > 0).",
            "  4. Conversion identity re-checks (e.g. impellerRadius == diameter/2; "
            "innerThickness == 2.28/innerChord*100).",
            "Approves or rejects with REVISE rationale via routing tool.",
            "(Optional in v5) DC_INSPECTOR_ENABLED toggles this stage off entirely.",
        ),
        "DC Input Inspector (DCII)",
        b(
            "Y",
            "Own self.messages persists across hand-offs.",
            "Image bytes stripped at on_operation_end.",
        ),
        b(
            "Hand-off from DC Input Creator",
            "parameters.json (current attempt)",
            "extracted_inputs.txt",
            "If present: Planner notes for DCII",
        ),
        b(
            "APPROVE or REVISE verdict, with detailed rationale:",
            "  - per-parameter range table",
            "  - user-locked-values match table",
            "  - hard-blocker check",
            "  - conversion-identity re-check",
            "  - notes on changes originating from upstream agents",
            "When REVISE: routes back to DCIC with the specific defect to address.",
        ),
        b(
            "Design Intent and Functional Requirements (Semantic)",
            "User-defined inputs (Floats and Ints) (Quantitative)",
            "User-defined inputs (2D files) (Quantitative) (Currently not implemented)",
            "User-defined inputs (3D files) (Quantitative) (Currently not implemented)",
            "User-defined inputs (image) (Semantic) (Currently not implemented)",
            "Input Parameters for Design Configurator (Floats and Ints) (Quantitative)",
            "Input Parameters for Design Configurator (2D files) (Quantitative) "
            "(Currently not implemented)",
            "Input Parameters for Design Configurator (3D files) (Quantitative) "
            "(Currently not implemented)",
            "Input Parameters for Design Configurator (images) (Semantic) "
            "(Currently not implemented)",
            "Failure Log (Semantic) (errors related to specific design inputs "
            "specified in the database)",
            "Design Output Description",
            "Design Correctness (Semantic) (did design output satisfy functional "
            "requirements?)",
        ),
        b(
            "• Validation-rejection cases library (Semantic): past instances where "
            "DCII caught a real issue — pattern, parameter combo, what made it visible. "
            "Builds a richer rejection repertoire than the four built-in checks alone.",
            "",
            "• Engineering-blocker corpus (Semantic + Quantitative): combinations "
            "beyond 'thickness ≤ 0' that historically produced bad meshes "
            "(non-watertight, fragmented, gross self-intersections). Catches what "
            "range-check alone cannot.",
            "",
            "• DCIC-revision-quality history (Semantic): when DCIC was asked to "
            "change parameters in response to a DCOI verdict, did the new set still "
            "honour all extraction-side callouts? Designed to catch regressions like "
            "the lost 2.28 mm callout in the recent recovery cycle.",
        ),
    ),

    # ============== 5. DC Output Inspector (DCOI) ==============
    (
        b(
            "Inspects the generated mesh + 3 renders and judges design correctness "
            "vs. the user's intent.",
            "Procedure (HARD RULE):",
            "  1. Load the 3 renders FIRST (load_render_images) and form an INDEPENDENT "
            "visual judgement.",
            "  2. THEN load the comparison source per DCOI_COMPARISON_MODE:",
            "     mode 1: USER INPUTS only (forbids reading extracted_inputs.txt)",
            "     mode 2: EXTRACTION only (forbids loading raw user inputs)",
            "     mode 3: extraction PRIMARY, user inputs SECONDARY (default)",
            "Two-sided count rule: count features in both render and reference.",
            "Mesh-quality numerics (watertight / volume / degenerate-face count) only "
            "available when MESH_CHECKS=True.",
        ),
        "DC Output Inspector (DCOI)",
        b(
            "Y",
            "Own self.messages persists across hand-offs.",
            "Image bytes stripped at on_operation_end (the agent re-loads renders "
            "via load_render_images on each turn it needs them).",
            "Buffer-and-flush ensures parallel image-loading + non-image tool calls "
            "produce a valid message history.",
        ),
        b(
            "Hand-off from Tool Caller (mesh path + 3 render paths)",
            "Render images (load_render_images)",
            "Comparison source per mode 1/2/3:",
            "  user_query.txt + paired reference image(s) + notes (mode 1, 3)",
            "  extracted_inputs.txt (mode 2, 3)",
            "Optional: deterministic mesh-quality metrics from render_and_check_mesh",
            "If present: planner notes for DCOI",
        ),
        b(
            "Structured verdict text:",
            "  - COMPARISON-SOURCE CLAIMS CHECKED (per-claim agreement / disagreement)",
            "  - GEOMETRY ANALYSIS",
            "  - DEFECTS",
            "  - DESIGN INTENT COMPLIANCE",
            "  - RECOMMENDATION (APPROVE or REVISE with corrective guidance for the "
            "next iteration)",
        ),
        b(
            "Design Intent and Functional Requirements (Semantic)",
            "Design output Files",
            "Design output renders",
            "Design Output Description",
            "Design Correctness (Semantic) (did design output satisfy functional "
            "requirements?)",
        ),
        b(
            "• Visual-comparison verdict library (Semantic + image): approved/rejected "
            "design renders paired with their reference images and verdict reasoning. "
            "Becomes a learning-to-judge corpus for image-match runs.",
            "",
            "• Defect catalog (Semantic + image): labelled images of detached blade "
            "tips, holes, spikes, fragmentation, non-watertight artifacts. Sharpens "
            "defect detection beyond the live render-resolution limits.",
            "",
            "• Cross-attempt comparison library (Semantic + image): past cases of "
            "'across N candidates, candidate X was the closest match because of "
            "feature Y' — directly serves the iterative best-fit workflow the user "
            "explicitly requests in v5 sessions.",
        ),
    ),

    # ============== 6. Tool Caller ==============
    (
        b(
            "Has the complete set of mesh-generation + rendering tools.",
            "Workflow per cycle:",
            "  1. read_parameters(parameters.json)",
            "  2. generate_propeller_mesh(...)  — RhinoCompute via .gh definition; "
            "writes propeller_mesh.obj, plus propeller_mesh_components.obj sidecar "
            "when MeshFinal is the primary path.",
            "  3. render_and_check_mesh(...)  — three PNGs + (when MESH_CHECKS=True) "
            "watertight / volume / degenerate-face metrics.",
            "Render-and-check backend is selectable per session via RENDER_LIBRARY in "
            "workflow_settings (trimesh OR pyvista).",
            "Hand-off to DCOI with absolute mesh + render paths.",
        ),
        "Tool Caller",
        b(
            "Y",
            "Own self.messages persists across hand-offs (was 'N' in original "
            "table — corrected for v5).",
        ),
        b(
            "Hand-off from DC Input Inspector",
            "Current attempt: <attempt_dir>",
            "Parameters file: <attempt_dir>/parameters.json",
        ),
        b(
            "Inside the attempt folder:",
            "  - propeller_mesh.obj",
            "  - propeller_mesh_components.obj (sidecar, written silently when "
            "MeshFinal was primary)",
            "  - render_isometric.png / render_top.png / render_side.png",
            "Optional metrics line: watertightness, volume, degenerate-face count "
            "(only when MESH_CHECKS=True)",
            "Hand-off to DC Output Inspector with all absolute paths verbatim.",
        ),
        b(
            "(none in original)",
        ),
        b(
            "• RhinoCompute error / timeout history (Semantic): past tool-call "
            "failures, what triggered them (network blip, parameter combo), and how "
            "they were recovered. Helps Tool Caller distinguish transient from "
            "terminal failures and decide retry vs escalate.",
            "",
            "• MeshFinal vs fallback diagnostic library (Semantic + Quantitative): "
            "which parameter combinations historically forced the fallback "
            "(per-component) path vs the welded MeshFinal. Lets the Tool Caller "
            "pre-warn downstream agents about likely watertightness issues.",
            "",
            "• Parameter-set → mesh-quality outcome history (Quantitative): past "
            "parameter sets paired with their final watertight / volume / "
            "degenerate-face numerics. Especially valuable when MESH_CHECKS=False "
            "this session — surrogate mesh-quality estimate from history.",
        ),
    ),

    # ============== 7. Context Pruner ==============
    (
        b(
            "(BUILT BUT NOT WIRED in v5 — agents/shared/context_pruner.py exists but "
            "no agent currently calls it.)",
            "Designed to: receive an agent's message history, then",
            "  - remove superfluous messages (old image renders no longer relevant, "
            "user messages from a different past request, stale tool-call messages)",
            "  - summarise non-critical messages (e.g. 'attempts when trying to fix a "
            "design' → one summary message; old visual renders → their qualitative "
            "and/or quantitative descriptions)",
            "Once wired, can be called by Receptionist, Planner, or Orchestrator on "
            "themselves OR on any other agent's history.",
        ),
        "Context Pruner (NOT WIRED)",
        b(
            "Y (when invoked)",
            "Receives the target agent's message history as input.",
        ),
        b(
            "Messages history for an agent (passed by the calling agent)",
        ),
        b(
            "Pruned messages history (returned to the calling agent for replacement)",
        ),
        b(
            "Problem formulation (Semantic)",
            "Solution to the problem (Semantic)",
            "Failure Log (Semantic)",
            "Design Intent and Functional Requirements (Semantic)",
            "Plan in Steps (Semantic)",
            "(same access as a Planner)",
        ),
        b(
            "• Pruning-rule library (Semantic): once wired, past examples of 'this "
            "kind of message was pruned this way' — keep / summarise / drop decisions "
            "with rationale. Builds a heuristic repertoire so the pruner doesn't "
            "rediscover policy from scratch each time.",
            "",
            "• Image-to-text replacement examples (Semantic): when an old render "
            "was replaced by its qualitative description in a prior run, what wording "
            "did the description use successfully? Avoids losing key visual info "
            "during summarisation.",
            "",
            "• Pruning-vs-failure correlation (Semantic): cases where "
            "over-aggressive pruning lost critical context and led to a downstream "
            "failure. Forces the pruner to err on the side of retention when the cost "
            "of dropping is uncertain.",
        ),
    ),

    # ============== 8. Orchestrator ==============
    (
        b(
            "Central coordinator and dispatch loop. Originates nothing; relays and "
            "shapes context.",
            "  - Receives every agent's hand-off message.",
            "  - Routes via call_<agent> tools (one per agent).",
            "  - On agent escalation: usually calls the Planner for the recovery "
            "directive, then routes onward.",
            "  - Final user-facing wrap-up via Receptionist when the cycle approves.",
            "Toggle CHAIN_ACCESS: when True, sees inter-agent messages exchanged "
            "while it was waiting (diagnostic-rich, expensive in tokens).",
            "v5 hardening: should distinguish design-side blocks from tool-side "
            "blocks; resists rubber-stamping a misframed escalation (e.g. DCIC's "
            "'tool wrapper is broken' framing of its own argument-shape bug).",
        ),
        "Orchestrator",
        b(
            "Y",
            "Own self.messages persists across hand-offs.",
            "Includes own system prompt as the first message.",
        ),
        b(
            "Response from any other agent",
            "Optional: chain access — every inter-agent message exchanged while "
            "waiting (when CHAIN_ACCESS=True)",
        ),
        b(
            "Calls the necessary agent next, with a message saying what to do and "
            "what to return.",
            "On approval: routes the final wrap-up to the Receptionist for the "
            "user-facing summary.",
        ),
        b(
            "Problem formulation (Semantic)",
            "Solution to the problem (Semantic)",
            "Plan in Steps (also showing which tools to call and what to check) "
            "(Semantic)",
        ),
        b(
            "• Routing-pattern library (Semantic): past cases of 'agent X "
            "escalates with framing Y → next routing was Z, and it worked / didn't.' "
            "Helps the Orchestrator avoid blindly trusting an agent's own framing of "
            "its block.",
            "",
            "• Escalation-misframing catalog (Semantic): cases where an agent "
            "blamed a tool wrapper / interface / model / something external when the "
            "actual cause was its own misuse (DCIC argument-shape, UII parenthetical "
            "noise). Guides skeptical re-routing.",
            "",
            "• Successful end-to-end session-shape library (Semantic): condensed "
            "routes (Receptionist → Planner → UII → DCIC → DCII → Tool Caller → DCOI "
            "→ Receptionist) of approved sessions, so the Orchestrator has a "
            "reference template for normal-flow vs anomaly recognition.",
        ),
    ),

    # ============== 9. Receptionist ==============
    (
        b(
            "Two entry points sharing one prompt:",
            "  - validate_input(input_dir): incoming user message — validates user "
            "inputs (image+note pairing case-insensitive by stem; basic norms; "
            "viability checks for any numeric values), then forwards to the "
            "Orchestrator OR replies directly to the user.",
            "  - format_outgoing(system_result): outgoing system message — formats "
            "agent output as friendly user-facing text (clarifications, status, "
            "success, failure, prompts for human feedback).",
            "Does NOT load image bytes — only reads notes via read_image_notes.",
            "Image-pairing gate: every image must have a paired _note.txt; orphans "
            "block forwarding.",
            "Ten-step budget; reply-direct path on confused / orphan inputs.",
        ),
        "Receptionist",
        b(
            "Y",
            "Own self.messages persists across hand-offs.",
            "Includes own system prompt as the first message.",
            "No image bytes ever — only path-text + notes.",
        ),
        b(
            "Option 1 (Situation A): receives user input as files stored in inputs/",
            "Option 2 (Situation B): receives a call from the Orchestrator with the "
            "system-side message to relay back to the user",
        ),
        b(
            "Option 1 outcomes:",
            "  1. If inputs are OK → notify Orchestrator that acceptable user inputs "
            "have been provided + where they're stored.",
            "  2. If inputs are not OK → reply to the user asking for correct inputs, "
            "explaining the rules.",
            "Option 2 outcome:",
            "  Reply to the user with whatever the Orchestrator requires "
            "(clarifications, status, follow-up question, etc.).",
        ),
        b(
            "Problem formulation (Semantic)",
            "Solution to the problem (Semantic)",
        ),
        b(
            "• User-clarification dialogue history (Semantic): past first-messages "
            "that needed clarification, what the Receptionist asked, and the user's "
            "eventual unblock. Helps anticipate the shape of future clarification "
            "loops.",
            "",
            "• Image-pairing failure examples (Semantic): off-topic notes, malformed "
            "pairings, orphan images — what was rejected, what was accepted with "
            "caveats, what slipped through. Tightens the gate over time.",
            "",
            "• Successful user-facing summary templates (Semantic): past wrap-up "
            "messages that resulted in user satisfaction or further engagement vs "
            "ones that led to confusion. Builds a friendlier, more actionable output "
            "style.",
        ),
    ),
]


# ---------------------------------------------------------------------------
# Build the workbook
# ---------------------------------------------------------------------------


def build_workbook(out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Agents (v5)"

    # ----- Header row -----
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFFFF")
    header_fill = PatternFill("solid", start_color="FF1F4E78")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # ----- Data rows -----
    body_font = Font(name="Arial", size=10)
    agent_font = Font(name="Arial", size=10, bold=True)
    body_align = Alignment(vertical="top", wrap_text=True)
    border_style = Side(border_style="thin", color="FFAAAAAA")
    cell_border = Border(
        top=border_style,
        bottom=border_style,
        left=border_style,
        right=border_style,
    )

    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = agent_font if col_idx == 2 else body_font
            c.alignment = body_align
            c.border = cell_border

    # ----- Column widths and row heights -----
    widths = {
        "A": 60,   # Goal of task/step
        "B": 22,   # Agent / Sub-Agent
        "C": 22,   # Has Message history
        "D": 40,   # Inputs
        "E": 45,   # Outputs
        "F": 50,   # Existing DB elements
        "G": 65,   # New RAG suggestions
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    # Reasonable row heights — tall enough for the multi-line content but
    # left as auto so Excel can re-flow if the user shrinks columns.
    ws.row_dimensions[1].height = 45
    for row_idx in range(2, len(ROWS) + 2):
        ws.row_dimensions[row_idx].height = 360

    # Freeze the header row + the agent-name column for easy scanning
    ws.freeze_panes = "C2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    build_workbook(OUT_PATH)
